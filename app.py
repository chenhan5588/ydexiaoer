import os
import io
import json
import uuid
import math
import base64
import statistics
import zipfile
import tempfile
import shutil
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from PIL import Image, ImageFilter, ImageStat, ImageOps, ImageEnhance
from pathlib import Path

# Excel generation
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Thread pool for parallel image processing
_io_pool = ThreadPoolExecutor(max_workers=4, thread_name_prefix="img")

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "printai-studio-secret-2024")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "printai2024")

# ---------------------------------------------------------------------------
# Login protection
# ---------------------------------------------------------------------------
from functools import wraps


def login_required(f):
    """Decorator: redirect to login if not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized", "redirect": "/login"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

PRINT_DPI = 300
CM_TO_INCH = 1 / 2.54

# Default: 30cm x 30cm (t-shirt transfer standard)
DEFAULT_PRINT_CM = 30


# Serve outputs directory for direct image access
@app.route("/outputs/<path:filename>")
@login_required
def serve_output(filename):
    return send_file(OUTPUT_DIR / filename)


def check_edge_quality(gray_img):
    """Check if edges are clean/smooth vs jagged/fragmented.

    Uses local gradient consistency: at each edge pixel, compare the
    gradient direction with its neighbors. Clean edges have consistent
    local directions; jagged edges have rapidly changing directions.
    """
    w, h = gray_img.size
    sample_size = min(w, 1200)
    if w > sample_size:
        gray_img = gray_img.resize((sample_size, int(h * sample_size / w)), Image.LANCZOS)

    # Sobel-like edge detection
    kx = ImageFilter.Kernel((3, 3), [-1, 0, 1, -2, 0, 2, -1, 0, 1], scale=4, offset=128)
    ky = ImageFilter.Kernel((3, 3), [-1, -2, -1, 0, 0, 0, 1, 2, 1], scale=4, offset=128)

    gx = gray_img.filter(kx)
    gy = gray_img.filter(ky)

    w2, h2 = gray_img.size
    pixels_gx = list(gx.getdata())
    pixels_gy = list(gy.getdata())

    # Calculate gradient magnitude and direction at each pixel
    magnitudes = []
    edge_pixels = []  # (y, x, direction)

    # First pass: find strong edge pixels
    edge_mask = [False] * (w2 * h2)
    for y in range(1, h2 - 1):
        for x in range(1, w2 - 1):
            idx = y * w2 + x
            dx = pixels_gx[idx] - 128
            dy = pixels_gy[idx] - 128
            mag = math.sqrt(dx * dx + dy * dy)
            magnitudes.append(mag)
            if mag > 12:  # strong enough edge
                dir_val = math.atan2(dy, dx)
                edge_pixels.append((y, x, dir_val))
                edge_mask[idx] = True

    if len(edge_pixels) < 50:
        return 0.0, "pass", "边缘柔和，无明显锯齿"

    # Second pass: for each edge pixel, check local direction consistency
    # Sample up to 500 edge pixels for performance
    import random
    sample_size = min(500, len(edge_pixels))
    sampled = random.sample(edge_pixels, sample_size)

    consistency_scores = []
    window = 3

    for cy, cx, main_dir in sampled:
        same_dir = 0
        total_neighbors = 0
        for dy in range(-window, window + 1):
            for dx in range(-window, window + 1):
                if dy == 0 and dx == 0:
                    continue
                ny, nx = cy + dy, cx + dx
                if 0 <= ny < h2 and 0 <= nx < w2:
                    nidx = ny * w2 + nx
                    if edge_mask[nidx]:
                        total_neighbors += 1
                        ndx = pixels_gx[nidx] - 128
                        ndy = pixels_gy[nidx] - 128
                        if ndx != 0 or ndy != 0:
                            ndir = math.atan2(ndy, ndx)
                            # Check if directions are similar (within 45 degrees)
                            diff = abs(ndir - main_dir)
                            if diff > math.pi:
                                diff = 2 * math.pi - diff
                            if diff < math.pi / 4:  # 45 degrees
                                same_dir += 1

        if total_neighbors >= 2:
            ratio = same_dir / total_neighbors
            consistency_scores.append(ratio)

    if not consistency_scores:
        return 0.0, "pass", "无法评估边缘"

    avg_consistency = statistics.mean(consistency_scores)
    # High consistency = clean edges, low = jagged
    # Invert so higher score = worse (for scoring consistency)
    jaggedness = 1.0 - avg_consistency

    if avg_consistency >= 0.6:
        return round(jaggedness, 3), "pass", "边缘清晰平滑"
    elif avg_consistency >= 0.35:
        return round(jaggedness, 3), "warn", "边缘存在轻微锯齿，建议描边检查"
    else:
        return round(jaggedness, 3), "fail", "边缘锯齿严重，需要修整"


def check_color_uniformity(rgb_img):
    """Check if colors are uniform and even across the image.

    Splits the image into a grid of patches and measures the standard
    deviation of brightness between patches. Low variance = uniform color,
    high variance = blotchy/uneven color.
    """
    w, h = rgb_img.size

    grid = 6
    patch_w = max(w // grid, 10)
    patch_h = max(h // grid, 10)

    patch_brightness = []
    for gy in range(grid):
        for gx in range(grid):
            x1 = gx * patch_w
            y1 = gy * patch_h
            x2 = min(x1 + patch_w, w)
            y2 = min(y1 + patch_h, h)
            if x2 - x1 < 5 or y2 - y1 < 5:
                continue
            patch = rgb_img.crop((x1, y1, x2, y2))
            stat = ImageStat.Stat(patch)
            avg = sum(stat.mean) / 3
            patch_brightness.append(avg)

    if len(patch_brightness) < 9:
        return 0.0, "pass", ""

    mean_brightness = statistics.mean(patch_brightness)
    if mean_brightness < 0.5:
        return 0.0, "pass", ""

    std_brightness = statistics.stdev(patch_brightness)
    cv = std_brightness / mean_brightness

    if cv < 0.06:
        return round(cv, 4), "pass", "颜色均匀"
    elif cv < 0.15:
        return round(cv, 4), "warn", "颜色轻微不均匀，建议检查"
    else:
        return round(cv, 4), "fail", "颜色不均匀，有坑洼或色差"


def analyze_image(img: Image.Image, filename: str, fast_mode: bool = False) -> dict:
    """Analyze image quality against t-shirt heat transfer standards.
    
    Args:
        fast_mode: Skip preview generation and heavy edge analysis for batch processing.
    """
    w, h = img.size

    # ---- Basic info ----
    total_pixels = w * h
    megapixels = round(total_pixels / 1_000_000, 2)
    mode = img.mode
    has_alpha = mode == "RGBA"
    mode_is_ok = mode in ("RGB", "RGBA")

    ext = Path(filename).suffix.lower()
    is_png = ext == ".png"

    # ---- 1. Resolution (for 30x30cm @ 300dpi -> need 3543px) ----
    eff_dpi = min(
        round(w / (DEFAULT_PRINT_CM * CM_TO_INCH)),
        round(h / (DEFAULT_PRINT_CM * CM_TO_INCH)),
    )

    # ---- 2. Sharpness (Laplacian variance) ----
    gray = img.convert("L")
    sample_w = min(w, 800)
    sample_h = int(h * sample_w / w) if w > 0 else sample_w
    gray_small = gray.resize((sample_w, sample_h), Image.LANCZOS)

    laplacian = gray_small.filter(ImageFilter.Kernel(
        (3, 3), [-1, -1, -1, -1, 8, -1, -1, -1, -1], scale=1, offset=0
    ))
    stat = ImageStat.Stat(laplacian)
    sharpness = stat.var[0]

    # ---- 3. Noise estimation ----
    blurred = gray_small.filter(ImageFilter.GaussianBlur(radius=2))
    diff_total = 0
    pixels = gray_small.width * gray_small.height
    for i, (orig, blur) in enumerate(zip(gray_small.getdata(), blurred.getdata())):
        diff_total += abs(orig - blur)
    noise_level = diff_total / pixels

    # ---- 4. Contrast ----
    contrast_stat = ImageStat.Stat(gray_small)
    contrast = contrast_stat.stddev[0] if contrast_stat.stddev else 0

    # ---- 5. Color uniformity (NEW) ----
    if mode in ("RGB", "RGBA"):
        rgb = img.convert("RGB")
        color_cv, color_grade, color_msg = check_color_uniformity(rgb)
    else:
        color_cv, color_grade, color_msg = 0.0, "pass", ""

    # ---- 6. Edge quality / jaggedness (skip in fast mode - very slow) ----
    if fast_mode:
        edge_cv, edge_grade, edge_msg = 0.0, "pass", ""
    else:
        edge_cv, edge_grade, edge_msg = check_edge_quality(gray)

    # ---- Scoring ----
    scores = {}
    issues = []

    # Resolution
    if eff_dpi >= 300:
        scores["resolution"] = "pass"
    elif eff_dpi >= 200:
        scores["resolution"] = "warn"
        issues.append(f"分辨率偏低 ({eff_dpi}dpi), 建议至少 300dpi (30cm需3543px)")
    else:
        scores["resolution"] = "fail"
        issues.append(f"分辨率严重不足 ({eff_dpi}dpi), 需要客户提供更高清图 (30cm需3543px)")

    # Sharpness
    if sharpness >= 200:
        scores["sharpness"] = "pass"
    elif sharpness >= 80:
        scores["sharpness"] = "warn"
        issues.append("图片轻微模糊, 锐化后可改善")
    else:
        scores["sharpness"] = "fail"
        issues.append("图片严重模糊, 不建议使用")

    # Noise
    if noise_level <= 8:
        scores["noise"] = "pass"
    elif noise_level <= 18:
        scores["noise"] = "warn"
        issues.append("图片有噪点, 降噪后可改善")
    else:
        scores["noise"] = "fail"
        issues.append("噪点严重, 烫画效果差")

    # Contrast
    if contrast >= 35:
        scores["contrast"] = "pass"
    elif contrast >= 20:
        scores["contrast"] = "warn"
        issues.append("对比度偏低, 烫画可能发灰")
    else:
        scores["contrast"] = "fail"
        issues.append("对比度过低, 印刷不清晰")

    # Color uniformity (NEW)
    scores["color_uniformity"] = color_grade
    if color_grade == "warn":
        issues.append(f"颜色轻微不均匀 ({color_cv:.3f}), 建议检查色差")
    elif color_grade == "fail":
        issues.append(f"颜色不均匀 ({color_cv:.3f}), 存在坑洼或色块差异")

    # Edge quality (NEW)
    scores["edge_quality"] = edge_grade
    if edge_grade == "warn":
        issues.append(f"边缘有锯齿 ({edge_cv:.2f}), 建议描边检查")
    elif edge_grade == "fail":
        issues.append(f"边缘锯齿严重 ({edge_cv:.2f}), 需要修整")

    # Format check (NEW)
    if not is_png:
        scores["format"] = "warn"
        issues.append("建议转换为 PNG 透明底格式")
    elif has_alpha:
        scores["format"] = "pass"
    else:
        scores["format"] = "warn"
        issues.append("PNG 格式正确，但缺少透明底，建议抠图后使用")

    if not has_alpha and not is_png:
        scores["format"] = "fail"
        issues.append("格式不符合要求：需要 PNG 透明底 (30x30cm, 300dpi)")

    # ---- Overall verdict ----
    fails = sum(1 for v in scores.values() if v == "fail")
    warns = sum(1 for v in scores.values() if v == "warn")

    if fails >= 2:
        verdict = "reject"
        verdict_label = "不可用"
    elif fails >= 1 or warns >= 3:
        verdict = "fix"
        verdict_label = "需修改"
    else:
        verdict = "pass"
        verdict_label = "可直接用"

    # Overall quality score 0-100
    quality_score = 100
    quality_score -= fails * 25
    quality_score -= warns * 10
    if megapixels < 1:
        quality_score -= 15
    if eff_dpi < 150:
        quality_score -= 20
    elif eff_dpi < 300:
        quality_score -= 10
    quality_score = max(0, min(100, quality_score))

    result = {
        "filename": filename,
        "width": w,
        "height": h,
        "megapixels": megapixels,
        "eff_dpi": eff_dpi,
        "sharpness": round(sharpness, 1),
        "noise_level": round(noise_level, 1),
        "contrast": round(contrast, 1),
        "color_uniformity": color_cv,
        "edge_quality": edge_cv,
        "mode": mode,
        "has_alpha": has_alpha,
        "is_png": is_png,
        "scores": scores,
        "issues": issues,
        "verdict": verdict,
        "verdict_label": verdict_label,
        "quality_score": quality_score,
        "fast_mode": fast_mode,
    }

    # Only generate base64 preview for non-fast mode (single image screen)
    if not fast_mode:
        preview = img.copy()
        max_side = 400
        if max(preview.size) > max_side:
            preview.thumbnail((max_side, max_side), Image.LANCZOS)
        buf = io.BytesIO()
        if preview.mode == "RGBA":
            bg = Image.new("RGB", preview.size, (255, 255, 255))
            bg.paste(preview, mask=preview.split()[3])
            preview = bg
        preview.save(buf, format="JPEG", quality=75)
        buf.seek(0)
        preview_b64 = base64.b64encode(buf.read()).decode()
        result["preview"] = f"data:image/jpeg;base64,{preview_b64}"

    return result


# ============== AI 修复引擎 ==============

def repair_upscale(img: Image.Image, target_dpi: int = 300, target_cm: int = 30) -> Image.Image:
    """Upscale image to target DPI for print size using Lanczos resampling.
    
    Fits within target dimensions (maintains aspect ratio, smaller dimension hits target).
    This avoids creating excessively large images for wide/pano formats.
    """
    target_px = int(target_cm * CM_TO_INCH * target_dpi)
    w, h = img.size
    # Fit within target: smaller dimension scaled to target, larger capped
    scale = target_px / max(w, h)
    if scale <= 1.0:
        return img  # Already large enough, no upscale needed
    new_w, new_h = int(w * scale), int(h * scale)
    return img.resize((new_w, new_h), Image.LANCZOS)


def repair_denoise(img: Image.Image) -> Image.Image:
    """Reduce noise using bilateral-like approach: median + subtle gaussian."""
    # First pass: median filter to remove salt-and-pepper noise
    img = img.filter(ImageFilter.MedianFilter(size=3))
    # Second pass: very subtle gaussian to smooth grain
    img = img.filter(ImageFilter.GaussianBlur(radius=0.6))
    return img


def repair_sharpen(img: Image.Image) -> Image.Image:
    """Apply unsharp mask to improve clarity without over-sharpening."""
    return img.filter(ImageFilter.UnsharpMask(radius=1.5, percent=120, threshold=3))


def repair_contrast(img: Image.Image) -> Image.Image:
    """Auto-enhance contrast for better print vibrancy."""
    enhancer = ImageEnhance.Contrast(img)
    # Analyze current contrast and adjust
    gray = img.convert("L")
    stat = ImageStat.Stat(gray)
    current_std = stat.stddev[0] if stat.stddev else 35
    # Target contrast boost: lower current contrast = stronger boost
    factor = max(1.0, min(2.0, 70 / max(current_std, 1)))
    return enhancer.enhance(factor)


def repair_color_uniformity(img: Image.Image) -> Image.Image:
    """Improve color uniformity via localized histogram equalization."""
    if img.mode == "RGBA":
        rgb = img.convert("RGB")
        alpha = img.getchannel("A")
    else:
        rgb = img.convert("RGB")
        alpha = None

    # Use ImageOps equalize (histogram equalization) for better uniformity
    r, g, b = rgb.split()
    r = ImageOps.equalize(r)
    g = ImageOps.equalize(g)
    b = ImageOps.equalize(b)
    equalized = Image.merge("RGB", (r, g, b))

    # Blend with original to avoid over-processing (70% equalized, 30% original)
    result = Image.blend(rgb, equalized, alpha=0.7)

    if alpha:
        result = result.convert("RGBA")
        result.putalpha(alpha)

    return result


def repair_remove_background(img: Image.Image) -> Image.Image:
    """Remove light/white background and create transparent PNG.
    
    Uses a threshold-based approach: pixels close to white become transparent.
    Works well for designs/logos on white or light backgrounds.
    """
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGBA")

    if img.mode == "RGBA":
        # Already has alpha - check if background is white and make it transparent
        r, g, b, a = img.split()
        rgb = Image.merge("RGB", (r, g, b))
    else:
        rgb = img.convert("RGB")
        a = Image.new("L", img.size, 255)

    pixels_rgb = list(rgb.getdata())
    pixels_a = list(a.getdata()) if a else [255] * len(pixels_rgb)

    new_alpha = []
    threshold = 200  # Brightness threshold for "white"
    for i, (r_val, g_val, b_val) in enumerate(pixels_rgb):
        brightness = (r_val + g_val + b_val) / 3
        if brightness > threshold:
            # Near-white pixel: make transparent proportionally
            opacity = max(0, int(255 * (brightness - threshold) / 55))
            new_alpha.append(min(pixels_a[i], opacity))
        else:
            new_alpha.append(pixels_a[i])

    result = rgb.convert("RGBA")
    result.putalpha(Image.new("L", img.size))
    result.putalpha(Image.new("L", img.size))
    # Rebuild alpha channel
    alpha_channel = Image.new("L", img.size)
    alpha_channel.putdata(new_alpha)
    result.putalpha(alpha_channel)

    return result


def repair_edge_smooth(img: Image.Image) -> Image.Image:
    """Smooth jagged edges using anti-aliasing approach.
    
    Downscale slightly then upscale back with Lanczos for natural edge smoothing.
    """
    w, h = img.size
    # Downscale to 80%, then back up
    small_w, small_h = int(w * 0.8), int(h * 0.8)
    small = img.resize((small_w, small_h), Image.LANCZOS)
    # Apply subtle blur at small size
    small = small.filter(ImageFilter.GaussianBlur(radius=0.5))
    # Upscale back
    result = small.resize((w, h), Image.LANCZOS)

    # Preserve alpha if present
    if img.mode == "RGBA":
        alpha = img.getchannel("A")
        result = result.convert("RGBA")
        result.putalpha(alpha)

    return result


def repair_image(img: Image.Image, scores: dict, issues: list) -> tuple:
    """Apply all relevant repairs based on detection results.
    
    Returns (repaired_image, applied_fixes_list).
    """
    applied = []
    repaired = img.copy()

    # Apply fixes: upscale first (lightweight ops before massive resize)
    fix_map = [
        ("denoise", scores.get("noise"), repair_denoise, "AI降噪"),
        ("sharpen", scores.get("sharpness"), repair_sharpen, "智能锐化"),
        ("contrast", scores.get("contrast"), repair_contrast, "对比度增强"),
        ("upscale", scores.get("resolution"), lambda i: repair_upscale(i), "超分辨率放大至300dpi"),
        ("color", scores.get("color_uniformity"), repair_color_uniformity, "颜色均匀化"),
        ("edge", scores.get("edge_quality"), repair_edge_smooth, "边���平滑"),
        ("background", scores.get("format"), repair_remove_background, "去背景→透明底PNG"),
    ]

    for key, grade, fix_fn, label in fix_map:
        if grade and grade in ("warn", "fail"):
            try:
                repaired = fix_fn(repaired)
                applied.append(label)
            except Exception as e:
                print(f"Repair {label} failed: {e}")

    # Always convert to PNG if not already
    if repaired.mode not in ("RGB", "RGBA"):
        repaired = repaired.convert("RGBA")

    return repaired, applied


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login_page():
    """Login page. GET shows form, POST verifies password."""
    error = None
    if request.method == "POST":
        if request.form.get("password") == APP_PASSWORD:
            session["logged_in"] = True
            session.permanent = True
            return redirect(url_for("index"))
        error = "密码错误，请重试"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/api/screen", methods=["POST"])
@login_required
def screen():
    files = request.files.getlist("images")
    if not files:
        return jsonify({"error": "请上传图片"}), 400

    target_w = float(request.form.get("printWidth", DEFAULT_PRINT_CM))
    target_h = float(request.form.get("printHeight", DEFAULT_PRINT_CM))

    results = []
    for f in files:
        if f.filename == "":
            continue
        ext = Path(f.filename).suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff", ".tif"):
            results.append({
                "filename": f.filename,
                "error": f"不支持的格式: {ext}",
            })
            continue

        try:
            img = Image.open(f.stream)
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            result = analyze_image(img, f.filename)

            # Recalculate DPI for target size
            eff_dpi_w = round(result["width"] / (target_w * CM_TO_INCH))
            eff_dpi_h = round(result["height"] / (target_h * CM_TO_INCH))
            result["eff_dpi"] = min(eff_dpi_w, eff_dpi_h)
            result["eff_dpi_w"] = eff_dpi_w
            result["eff_dpi_h"] = eff_dpi_h
            result["print_width_cm"] = target_w
            result["print_height_cm"] = target_h

            # Re-score resolution
            if result["eff_dpi"] >= 300:
                result["scores"]["resolution"] = "pass"
            elif result["eff_dpi"] >= 200:
                result["scores"]["resolution"] = "warn"
            else:
                result["scores"]["resolution"] = "fail"

            # Recalculate verdict
            fails = sum(1 for v in result["scores"].values() if v == "fail")
            warns = sum(1 for v in result["scores"].values() if v == "warn")
            if fails >= 2:
                result["verdict"] = "reject"
                result["verdict_label"] = "不可用"
            elif fails >= 1 or warns >= 3:
                result["verdict"] = "fix"
                result["verdict_label"] = "需修改"
            else:
                result["verdict"] = "pass"
                result["verdict_label"] = "可直接用"

            quality_score = 100 - fails * 25 - warns * 10
            if result["megapixels"] < 1:
                quality_score -= 15
            if result["eff_dpi"] < 150:
                quality_score -= 20
            elif result["eff_dpi"] < 300:
                quality_score -= 10
            result["quality_score"] = max(0, min(100, quality_score))

            results.append(result)
        except Exception as e:
            results.append({
                "filename": f.filename,
                "error": f"读取失败: {str(e)}",
            })

    pass_count = sum(1 for r in results if r.get("verdict") == "pass")
    fix_count = sum(1 for r in results if r.get("verdict") == "fix")
    reject_count = sum(1 for r in results if r.get("verdict") == "reject")
    error_count = sum(1 for r in results if "error" in r)

    return jsonify({
        "results": results,
        "summary": {
            "total": len(results),
            "pass": pass_count,
            "fix": fix_count,
            "reject": reject_count,
            "error": error_count,
        },
        "print_size": f"{target_w}x{target_h}cm",
    })


@app.route("/api/repair", methods=["POST"])
@login_required
def repair():
    """AI repair endpoint: accepts image + metadata, returns fixed image.
    
    Can repair a single image or re-process from the analysis results.
    """
    file = request.files.get("image")
    if not file or file.filename == "":
        return jsonify({"error": "请上传图片"}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff", ".tif"):
        return jsonify({"error": f"不支持的格式: {ext}"}), 400

    try:
        img = Image.open(file.stream)
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGBA")

        # Run analysis to get scores
        analysis = analyze_image(img, file.filename)

        # Apply repairs
        repaired, applied_fixes = repair_image(img, analysis["scores"], analysis["issues"])

        # Generate preview
        preview_img = repaired.copy()
        max_side = 400
        if max(preview_img.size) > max_side:
            preview_img.thumbnail((max_side, max_side), Image.LANCZOS)
        buf = io.BytesIO()
        preview_img.save(buf, format="PNG")
        buf.seek(0)
        preview_b64 = base64.b64encode(buf.read()).decode()

        # Save full-resolution repaired image
        file_id = uuid.uuid4().hex[:12]
        output_path = OUTPUT_DIR / f"repaired_{file_id}.png"
        repaired.save(output_path, format="PNG")

        # Re-analyze repaired image
        repaired_check = analyze_image(repaired, f"repaired_{Path(file.filename).stem}.png")

        return jsonify({
            "filename": file.filename,
            "repaired_filename": f"repaired_{Path(file.filename).stem}.png",
            "file_id": file_id,
            "preview": f"data:image/png;base64,{preview_b64}",
            "applied_fixes": applied_fixes,
            "original_score": analysis["quality_score"],
            "repaired_score": repaired_check["quality_score"],
            "original_verdict": analysis["verdict_label"],
            "repaired_verdict": repaired_check["verdict_label"],
            "repaired_issues": repaired_check["issues"],
            "download_url": f"/api/download/{file_id}",
        })

    except Exception as e:
        return jsonify({"error": f"修复失败: {str(e)}"}), 500


@app.route("/api/repair-batch", methods=["POST"])
@login_required
def repair_batch():
    """Batch repair: accepts multiple images, returns all repaired results."""
    files = request.files.getlist("images")
    if not files:
        return jsonify({"error": "请上传图片"}), 400

    results = []
    for f in files:
        if f.filename == "":
            continue
        ext = Path(f.filename).suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff", ".tif"):
            results.append({"filename": f.filename, "error": f"不支持的格式: {ext}"})
            continue

        try:
            img = Image.open(f.stream)
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")

            analysis = analyze_image(img, f.filename)
            repaired, applied_fixes = repair_image(img, analysis["scores"], analysis["issues"])

            file_id = uuid.uuid4().hex[:12]
            output_path = OUTPUT_DIR / f"repaired_{file_id}.png"
            repaired.save(output_path, format="PNG")

            # Preview
            preview_img = repaired.copy()
            max_side = 300
            if max(preview_img.size) > max_side:
                preview_img.thumbnail((max_side, max_side), Image.LANCZOS)
            buf = io.BytesIO()
            preview_img.save(buf, format="PNG")
            buf.seek(0)
            preview_b64 = base64.b64encode(buf.read()).decode()

            results.append({
                "filename": f.filename,
                "repaired_filename": f"repaired_{Path(f.filename).stem}.png",
                "file_id": file_id,
                "preview": f"data:image/png;base64,{preview_b64}",
                "applied_fixes": applied_fixes,
                "original_score": analysis["quality_score"],
                "download_url": f"/api/download/{file_id}",
            })
        except Exception as e:
            results.append({"filename": f.filename, "error": f"修复失败: {str(e)}"})

    return jsonify({"results": results})


@app.route("/api/download/<file_id>")
@login_required
def download_fixed(file_id):
    """Download a repaired image by its file_id."""
    path = OUTPUT_DIR / f"repaired_{file_id}.png"
    if not path.exists():
        return jsonify({"error": "文件不存在或已过期"}), 404
    return send_file(path, mimetype="image/png", as_attachment=True,
                     download_name=f"repaired_{file_id}.png")


if __name__ == "__main__":
    app.run(debug=True, port=5051)


# ============== 订单 ZIP 批量解析 ==============

def extract_nested_zip(zip_path, extract_to):
    """Recursively extract nested zip files."""
    extracted = []
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(extract_to)
        for name in zf.namelist():
            full_path = os.path.join(extract_to, name)
            if name.lower().endswith('.zip') and os.path.isfile(full_path):
                sub_dir = os.path.join(extract_to, f"_nested_{uuid.uuid4().hex[:8]}")
                os.makedirs(sub_dir, exist_ok=True)
                extracted.extend(extract_nested_zip(full_path, sub_dir))
    extracted.append(extract_to)
    return extracted


def find_files(root_dir, extensions):
    """Find all files with given extensions under root_dir."""
    results = []
    for dirpath, _, filenames in os.walk(root_dir):
        for fn in filenames:
            if any(fn.lower().endswith(ext) for ext in extensions):
                results.append(os.path.join(dirpath, fn))
    return results


def parse_size_from_title(title):
    """Extract size from Amazon product title, e.g. '..., Schwarz, M' -> 'M'."""
    if not title:
        return ""
    # Common size patterns at end of title
    patterns = [
        r',\s*([XSML]|XL|XXL|XXXL|2XL|3XL|4XL|5XL)\s*$',
        r',\s*(\d{2,3}[cm]?\s*(?:/\s*\d{2,3}[cm]?)?)\s*$',
        r'-([XSML]|XL|XXL|XXXL|2XL|3XL|4XL|5XL)\s*$',
    ]
    for pat in patterns:
        m = re.search(pat, title, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ""


def parse_size_from_sku(sku_or_filename):
    """Extract size from SKU like '1946VEST4-Black-M'."""
    if not sku_or_filename:
        return ""
    parts = sku_or_filename.replace('_', '-').split('-')
    for p in parts:
        p_upper = p.strip().upper()
        if p_upper in ('XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL', '2XL', '3XL', '4XL', '5XL'):
            return p.strip()
    return ""


def find_order_json(root_dir):
    """Find the order JSON file in extracted directory."""
    json_files = find_files(root_dir, ['.json'])
    # Prefer files that look like Amazon order data (have orderId)
    for jf in json_files:
        try:
            with open(jf, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if 'orderId' in data or 'customizationData' in data:
                return jf, data
        except Exception:
            continue
    return None, None


def find_images_in_order(root_dir, order_data):
    """Find snapshot (effect image) and customer upload images."""
    all_images = find_files(root_dir, ['.jpg', '.jpeg', '.png'])
    snapshots = []
    customer_images = []
    
    if not order_data:
        return snapshots, customer_images
    
    def _find_all(node, surface_label=""):
        """Recursively find all ImageCustomization nodes in the JSON tree."""
        if isinstance(node, dict):
            # Check if this is a surface with label
            if node.get('type') == 'PreviewContainerCustomization':
                surface_label = node.get('label', '')
            
            # Check for snapshot
            snap = node.get('snapshot', {})
            if isinstance(snap, dict) and snap.get('imageName'):
                snap_name = snap['imageName']
                for img_path in all_images:
                    if snap_name.lower() in img_path.lower():
                        snapshots.append(img_path)
                        break
            
            # Check for ImageCustomization
            if node.get('type') == 'ImageCustomization':
                img_info = node.get('image', {})
                img_name = img_info.get('imageName')
                if img_name:
                    for img_path in all_images:
                        if img_name.lower() in img_path.lower():
                            customer_images.append({
                                'path': img_path,
                                'buyer_filename': img_info.get('buyerFilename', ''),
                                'surface': surface_label,
                            })
                            break
            
            # Recurse into all values
            for v in node.values():
                _find_all(v, surface_label)
        elif isinstance(node, list):
            for item in node:
                _find_all(item, surface_label)
    
    try:
        _find_all(order_data)
    except Exception as e:
        print(f"Error finding images: {e}")
    
    return snapshots, customer_images


def process_order_zip(zip_path, mabang_order_id, do_repair=True):
    """Process a single order ZIP file. Returns order dict."""
    temp_dir = tempfile.mkdtemp(prefix="order_")
    try:
        # Extract nested zips
        extract_nested_zip(zip_path, temp_dir)
        
        # Find order JSON
        json_path, order_data = find_order_json(temp_dir)
        if not order_data:
            return {"error": "无法解析订单数据", "mabang_order_id": mabang_order_id}
        
        # Extract basic info
        order_id = order_data.get('orderId', '')
        asin = order_data.get('asin', '')
        title = order_data.get('title', '')
        quantity = order_data.get('quantity', 1)
        
        # Extract size
        size = parse_size_from_title(title) or parse_size_from_sku(mabang_order_id)
        
        # Find images
        snapshots, customer_images = find_images_in_order(temp_dir, order_data)
        
        # Save snapshots
        saved_snapshots = []
        for snap_path in snapshots:
            try:
                snap_img = Image.open(snap_path)
                snap_fid = uuid.uuid4().hex[:12]
                snap_save = OUTPUT_DIR / f"snapshot_{snap_fid}_{os.path.basename(snap_path)}"
                if snap_img.mode == 'RGBA':
                    snap_img.convert('RGB').save(snap_save, 'JPEG', quality=90)
                else:
                    snap_img.save(snap_save, 'JPEG', quality=90)
                saved_snapshots.append(str(snap_save))
            except Exception as e:
                print(f"Failed to save snapshot {snap_path}: {e}")
                saved_snapshots.append(snap_path)
        
        # ---- Parallel image processing ----
        analyzed_images = []
        
        def _process_one_image(ci):
            """Process a single customer image: load -> analyze -> repair -> save."""
            try:
                img = Image.open(ci['path'])
                if img.mode not in ('RGB', 'RGBA'):
                    img = img.convert('RGBA')
                
                # Fast analysis (skip preview + edge quality for speed)
                analysis = analyze_image(img, os.path.basename(ci['path']), fast_mode=True)
                analysis['surface'] = ci['surface']
                analysis['buyer_filename'] = ci['buyer_filename']
                
                # Save original for Excel access
                orig_fid = uuid.uuid4().hex[:12]
                orig_save_path = OUTPUT_DIR / f"orig_{orig_fid}_{os.path.basename(ci['path'])}"
                img.save(orig_save_path, format='PNG')
                analysis['saved_original_path'] = str(orig_save_path)
                
                # Repair if needed
                if do_repair and analysis['verdict'] != 'pass':
                    try:
                        repaired, fixes = repair_image(img, analysis['scores'], analysis['issues'])
                        fid = uuid.uuid4().hex[:12]
                        repaired_path = OUTPUT_DIR / f"repaired_{fid}.png"
                        repaired.save(repaired_path, format='PNG')
                        analysis['repaired_path'] = str(repaired_path)
                        analysis['applied_fixes'] = fixes
                        
                        # Fast re-analysis of repaired image
                        re_check = analyze_image(repaired, f"repaired_{os.path.basename(ci['path'])}", fast_mode=True)
                        analysis['repaired_score'] = re_check['quality_score']
                        analysis['repaired_verdict'] = re_check['verdict_label']
                    except Exception as e:
                        analysis['repair_error'] = str(e)
                
                return analysis
            except Exception as e:
                return {
                    'filename': os.path.basename(ci['path']),
                    'error': str(e),
                    'surface': ci.get('surface', ''),
                    'buyer_filename': ci.get('buyer_filename', ''),
                }
        
        # Process all images in parallel
        if customer_images:
            futures = [_io_pool.submit(_process_one_image, ci) for ci in customer_images]
            for future in as_completed(futures):
                analyzed_images.append(future.result())
        
        return {
            "mabang_order_id": mabang_order_id,
            "order_id": order_id,
            "asin": asin,
            "title": title,
            "size": size,
            "quantity": quantity,
            "snapshot_paths": saved_snapshots,
            "customer_images": analyzed_images,
        }
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.route("/api/batch-orders", methods=["POST"])
@login_required
def batch_orders():
    """Upload order ZIP files, parse and analyze all.
    
    Uses a two-phase approach:
    1. Extract all ZIPs and parse order data
    2. Flatten all images across orders into a single parallel processing pool
    """
    files = request.files.getlist("zips")
    if not files:
        return jsonify({"error": "请上传 ZIP 文件"}), 400
    
    do_repair = request.form.get("repair", "true").lower() == "true"
    
    # Phase 1: Extract all ZIPs and collect order metadata + image refs
    orders_meta = []  # list of {mabang_id, order_data, temp_zip_path, temp_dir}
    all_image_tasks = []  # list of (ci_dict, mabang_id_idx)
    
    for f in files:
        if f.filename == "" or not f.filename.lower().endswith('.zip'):
            continue
        
        mabang_id = Path(f.filename).stem
        temp_zip = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
        f.save(temp_zip.name)
        temp_zip.close()
        
        try:
            temp_dir = tempfile.mkdtemp(prefix="order_")
            extract_nested_zip(temp_zip.name, temp_dir)
            
            json_path, order_data = find_order_json(temp_dir)
            if not order_data:
                orders_meta.append({
                    "mabang_id": mabang_id,
                    "error": "无法解析订单数据",
                    "temp_dir": temp_dir,
                    "temp_zip": temp_zip.name,
                })
                continue
            
            order_id = order_data.get('orderId', '')
            asin = order_data.get('asin', '')
            title = order_data.get('title', '')
            quantity = order_data.get('quantity', 1)
            size = parse_size_from_title(title) or parse_size_from_sku(mabang_id)
            
            snapshots, customer_images = find_images_in_order(temp_dir, order_data)
            
            orders_meta.append({
                "mabang_id": mabang_id,
                "order_id": order_id,
                "asin": asin,
                "title": title,
                "size": size,
                "quantity": quantity,
                "snapshots": snapshots,
                "customer_images": customer_images,
                "temp_dir": temp_dir,
                "temp_zip": temp_zip.name,
                "error": None,
            })
        except Exception as e:
            orders_meta.append({
                "mabang_id": mabang_id,
                "error": str(e),
                "temp_dir": None,
                "temp_zip": temp_zip.name,
            })
    
    # Phase 2: Flatten all images across orders, process in parallel
    # Build a flat list of image processing tasks
    flat_tasks = []  # (image_info, meta_idx)
    for idx, meta in enumerate(orders_meta):
        if meta.get("error"):
            continue
        for ci in meta.get("customer_images", []):
            flat_tasks.append((ci, idx))
    
    # Process all images in one parallel batch
    image_results = {}  # meta_idx -> list of analysis results
    
    def _process_flat(ci, meta_idx):
        try:
            img = Image.open(ci['path'])
            if img.mode not in ('RGB', 'RGBA'):
                img = img.convert('RGBA')
            
            analysis = analyze_image(img, os.path.basename(ci['path']), fast_mode=True)
            analysis['surface'] = ci['surface']
            analysis['buyer_filename'] = ci['buyer_filename']
            
            # Save original
            orig_fid = uuid.uuid4().hex[:12]
            orig_save_path = OUTPUT_DIR / f"orig_{orig_fid}_{os.path.basename(ci['path'])}"
            img.save(orig_save_path, format='PNG')
            analysis['saved_original_path'] = str(orig_save_path)
            
            if do_repair and analysis['verdict'] != 'pass':
                try:
                    repaired, fixes = repair_image(img, analysis['scores'], analysis['issues'])
                    fid = uuid.uuid4().hex[:12]
                    repaired_path = OUTPUT_DIR / f"repaired_{fid}.png"
                    repaired.save(repaired_path, format='PNG')
                    analysis['repaired_path'] = str(repaired_path)
                    analysis['applied_fixes'] = fixes
                    
                    re_check = analyze_image(repaired, f"repaired_{os.path.basename(ci['path'])}", fast_mode=True)
                    analysis['repaired_score'] = re_check['quality_score']
                    analysis['repaired_verdict'] = re_check['verdict_label']
                except Exception as e:
                    analysis['repair_error'] = str(e)
            
            return meta_idx, analysis
        except Exception as e:
            return meta_idx, {
                'filename': os.path.basename(ci['path']),
                'error': str(e),
                'surface': ci.get('surface', ''),
                'buyer_filename': ci.get('buyer_filename', ''),
            }
    
    if flat_tasks:
        futures = [_io_pool.submit(_process_flat, ci, idx) for ci, idx in flat_tasks]
        for future in as_completed(futures):
            meta_idx, result = future.result()
            image_results.setdefault(meta_idx, []).append(result)
    
    # Phase 3: Assemble final results
    results = []
    errors = []
    
    for meta in orders_meta:
        if meta.get("error"):
            errors.append({"filename": meta["mabang_id"], "error": meta["error"]})
            # Cleanup
            if meta.get("temp_dir"):
                shutil.rmtree(meta["temp_dir"], ignore_errors=True)
            if meta.get("temp_zip"):
                os.unlink(meta["temp_zip"])
            continue
        
        # Save snapshots
        saved_snapshots = []
        for snap_path in meta.get("snapshots", []):
            try:
                snap_img = Image.open(snap_path)
                snap_fid = uuid.uuid4().hex[:12]
                snap_save = OUTPUT_DIR / f"snapshot_{snap_fid}_{os.path.basename(snap_path)}"
                if snap_img.mode == 'RGBA':
                    snap_img.convert('RGB').save(snap_save, 'JPEG', quality=90)
                else:
                    snap_img.save(snap_save, 'JPEG', quality=90)
                saved_snapshots.append(str(snap_save))
            except Exception as e:
                saved_snapshots.append(snap_path)
        
        results.append({
            "mabang_order_id": meta["mabang_id"],
            "order_id": meta["order_id"],
            "asin": meta["asin"],
            "title": meta["title"],
            "size": meta["size"],
            "quantity": meta["quantity"],
            "snapshot_paths": saved_snapshots,
            "customer_images": image_results.get(orders_meta.index(meta), []),
        })
        
        # Cleanup
        if meta.get("temp_dir"):
            shutil.rmtree(meta["temp_dir"], ignore_errors=True)
        if meta.get("temp_zip"):
            os.unlink(meta["temp_zip"])
    
    return jsonify({
        "orders": results,
        "errors": errors,
        "total": len(files),
        "success": len(results),
        "failed": len(errors),
    })


@app.route("/api/export-excel", methods=["POST"])
@login_required
def export_excel():
    """Generate Excel from parsed order data."""
    data = request.get_json()
    if not data or 'orders' not in data:
        return jsonify({"error": "缺少订单数据"}), 400
    
    orders = data['orders']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "订单汇总"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["马帮订单号", "订单编号", "产品图片（效果图）", "尺码", "数量", "特殊要求（来图）"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    row = 2
    img_height = 120  # pixels
    
    for order in orders:
        mabang_id = order.get('mabang_order_id', '')
        order_id = order.get('order_id', '')
        size = order.get('size', '')
        quantity = order.get('quantity', 1)
        
        ws.cell(row=row, column=1, value=mabang_id).border = thin_border
        ws.cell(row=row, column=2, value=order_id).border = thin_border
        ws.cell(row=row, column=4, value=size).border = thin_border
        ws.cell(row=row, column=5, value=quantity).border = thin_border
        
        # Product image (snapshot/effect image)
        snapshots = order.get('snapshot_paths', [])
        if snapshots and os.path.exists(snapshots[0]):
            try:
                xl_img = XLImage(snapshots[0])
                xl_img.width = 160
                xl_img.height = 120
                ws.add_image(xl_img, f"C{row}")
            except Exception as e:
                ws.cell(row=row, column=3, value=f"[图片: {e}]").border = thin_border
        
        # Special requirement (customer images)
        customer_images = order.get('customer_images', [])
        if customer_images:
            ci = customer_images[0]
            img_path = ci.get('repaired_path') or ci.get('saved_original_path')
            
            if img_path and os.path.exists(img_path):
                try:
                    xl_img2 = XLImage(img_path)
                    xl_img2.width = 160
                    xl_img2.height = 120
                    ws.add_image(xl_img2, f"F{row}")
                except Exception as e:
                    ws.cell(row=row, column=6, value=f"[图片: {e}]").border = thin_border
            else:
                issues_text = '; '.join(ci.get('issues', [])[:2])
                ws.cell(row=row, column=6, value=issues_text or "无来图").border = thin_border
        
        # Set row height
        ws.row_dimensions[row].height = img_height * 0.75  # convert to points
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 28
    
    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"订单汇总_{uuid.uuid4().hex[:8]}.xlsx"
    )
