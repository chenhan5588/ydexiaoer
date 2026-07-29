"""Excel export for batch order results.

Generates a formatted .xlsx with order summary including embedded product
images and customer upload images.
"""

import io
import uuid
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def export_excel_file(orders):
    """Generate Excel file from parsed order data.

    Args:
        orders: List of order dicts from batch processing.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "订单汇总"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["马帮订单号", "订单编号", "产品图片（效果图）", "尺码", "数量", "特殊要求（来图）"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    img_height = 120

    for order in orders:
        mabang_id = order.get('mabang_order_id', '')
        order_id = order.get('order_id', '')
        size = order.get('size', '')
        quantity = order.get('quantity', 1)

        ws.cell(row=row, column=1, value=mabang_id).border = thin_border
        ws.cell(row=row, column=2, value=order_id).border = thin_border
        ws.cell(row=row, column=4, value=size).border = thin_border
        ws.cell(row=row, column=5, value=quantity).border = thin_border

        # Product image (snapshot/effect image)
        snapshots = order.get('snapshot_paths', [])
        if snapshots and os.path.exists(snapshots[0]):
            try:
                xl_img = XLImage(snapshots[0])
                xl_img.width = 160
                xl_img.height = 120
                ws.add_image(xl_img, f"C{row}")
            except Exception as e:
                ws.cell(row=row, column=3, value=f"[图片: {e}]").border = thin_border

        # Customer images
        customer_images = order.get('customer_images', [])
        if customer_images:
            ci = customer_images[0]
            img_path = ci.get('repaired_path') or ci.get('saved_original_path')
            if img_path and os.path.exists(img_path):
                try:
                    xl_img2 = XLImage(img_path)
                    xl_img2.width = 160
                    xl_img2.height = 120
                    ws.add_image(xl_img2, f"F{row}")
                except Exception as e:
                    ws.cell(row=row, column=6, value=f"[图片: {e}]").border = thin_border
            else:
                issues_text = '; '.join(ci.get('issues', [])[:2])
                ws.cell(row=row, column=6, value=issues_text or "无来图").border = thin_border

        ws.row_dimensions[row].height = img_height * 0.75
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
